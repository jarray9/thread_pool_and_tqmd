#!/user/bin/python
# -*- encoding : utf-8 -*-

import openpyxl
from openpyxl import Workbook
import pandas as pd
from openpyxl.utils import get_column_letter

import const
from const import File


def init_DataFrame():
    return pd.DataFrame()


def init_report(file_name: str):
    """Create data of report for test."""
    wb = Workbook()
    sheet = wb.active
    sheet.append(File.header)
    for x in range(10):
        if x % 2 == 0:
            sheet.append(File.content1)
        else:
            sheet.append(File.content2)
    wb.save(filename=file_name)


def read_excel_to_dataframe(file_name, sort_by=File.header[1]):
    try:
        df = pd.read_excel(file_name)
        return df.sort_values(by=[sort_by])
    except FileNotFoundError:
        raise Exception(f"File is not exits :{file_name}")
    except Exception as e:
        raise Exception(f"Unexpected error occurred.{e.args}")


def read_excel_file(file_name):
    try:
        wb = openpyxl.load_workbook(file_name=file_name)
        if "sheet1" in wb.sheetnames:
            return wb[File.sheet_name]
        else:
            raise Exception(f"Expected sheet name does not exist. {file_name}")
    except FileNotFoundError:
        raise Exception(f"File is not exits :{file_name}")
    except Exception as e:
        raise Exception(f"Unexpected error occurred.{e.args}")


def merge_content(data_frame_list):
    """
    Concat multiple dataframe into 1
    """
    df = pd.concat(data_frame_list, ignore_index=True)
    new_df = df.drop(df[(df.res == const.status_ok) & (df.ステータス == const.status_done)].index)
    return new_df.reset_index(drop=True)
